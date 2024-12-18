"""utils module for yadisk interactions
"""
import csv
import os

from openpyxl import load_workbook

from yadisk_manager import DiskManager


CSV_DELIMITER = os.getenv('CSV_DELIMITER', ';')


def add_csv_to_table(csv_filepath, workbook, sheet_name='export', delimiter=CSV_DELIMITER):
    # delete existing sheet to rewrite
    if sheet_name in workbook.sheetnames:
        #workbook.remove(workbook[sheet_name])
        ws = workbook[sheet_name]
        ws.insert_rows(idx=0, amount=0)    # clear sheet instead removing, that can break formulas
    else:
        ws = workbook.create_sheet(sheet_name)
    
    with open(csv_filepath, encoding="utf-8") as f:
        reader = csv.reader(f, delimiter=delimiter)
        for row in reader:
            ws.append(row)


def write_sheet_to_file(yatoken, remote_path, csv_path, sheet_name='export'):
    disk_manager = DiskManager(yatoken=yatoken)

    # download file to filesystem
    local_path = disk_manager.download_file_from_disk(remote_path)

    # create openpyxl.Workbook from existing xlsx file
    wb = load_workbook(filename=local_path)

    # add csv to table as sheet 
    add_csv_to_table(csv_path, wb, sheet_name=sheet_name)

    # save openpyxl.Workbook to filesystem with same name
    wb.save(local_path)

    # download file to disk
    disk_manager.upload(local_path, remote_path)

    os.remove(local_path)
